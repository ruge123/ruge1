import openpyxl  # (导入openpyxl库)
import requests


# 读取测试用例
def read_data(filename, sheetname):
    wb = openpyxl.load_workbook(filename)
    sh = wb[sheetname]
    max_row = sh.max_row
    case_list = []

    for i in range(2, max_row + 1):
        dict1 = dict(
            case_id = sh.cell(row=i,column=1).value,
            url	=sh.cell(row=i ,column=5).value,
            data =sh.cell(row=i ,column=6).value,
            expect =sh.cell(row=i ,column=7).value)
        case_list.append(dict1)
    return case_list

# 发送接口请求
def api_fun(url,data):
    headers = {'X-Lemonban-Media-Type':'lemonban.v2', 'Content-Type':'application/json'}
    result = requests.post(url=url,json=data,headers=headers).json()
    return result

# 写入结果
def wirte_result(filename,sheetname,row,column,final_result):
    wb = openpyxl.load_workbook(filename)
    sh = wb[sheetname]
    sh.cell(row=row,column=column).value = final_result
    wb.save(filename)

# 执行接口测试用例
def execute_fun(filename,sheetname):
    cases = read_data(filename,sheetname)
    for case in cases:
        case_id = case['case_id']
        url = case['url']
        data = eval(case['data'])
        expect = eval(case['expect'])
        expect_code = expect['code']
        expect_msg = expect['msg']
        print(f'预期code为{expect_code}，msg为{expect_msg}')

        real_result = api_fun(url=url,data=data)
        # print(real_result)

        real_code = real_result['code']
        real_msg = real_result['msg']
        print(f'实际code为{real_code}，msg为{real_msg}')

        if real_code == expect_code and real_msg == expect_msg:
            print(f'第{case_id}条用例通过！')
            final_result ='Passed'
        else:
            print(f'第{case_id}条用例不通过')
            final_result = 'Failed'

        wirte_result(filename,sheetname,case_id+1,8,final_result) #

execute_fun('test_case_api.xlsx','login')


